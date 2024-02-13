from openpyxl import Workbook, load_workbook
wb = load_workbook('C:/Users/Vee/OneDrive - Goodlife Pharmacy/Desktop/MY PROJECTS/Randoms/Web Scraping/OUTPUT/Excel Attachments/206_Goodlife Shell Tegemeo/URGENT SPECIAL ORDER 8.2.2024.xlsx')
ws = wb.active
print(ws['A3'].value)