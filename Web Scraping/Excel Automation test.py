import os
from openpyxl import load_workbook

# Directory containing the Excel files
# home laptop
# base_dir = 'C:/Users/Vee/OneDrive - Goodlife Pharmacy/Desktop/MY PROJECTS/Randoms/Web Scraping/OUTPUT/Excel Attachments'

# work laptop
base_dir = 'C:/Users/VMUKITA/OneDrive - Goodlife Pharmacy/Desktop/MY PROJECTS/Randoms/Web Scraping/OUTPUT/Excel Attachments'

def find_code_column(directory):
    # Create a list to store the file paths
    file_paths = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                file_paths.append(file_path)

    # Sort the file paths based on their index in the folder name
    file_paths.sort(key=lambda x: int(os.path.basename(os.path.dirname(x)).split('_')[0]))

    for file_path in file_paths:
        try:
            # Load the workbook
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                # Check if the sheet contains a column named "code"
                sheet = wb[sheet_name]
                if any('code' in str(cell.value).lower() for row in sheet.iter_rows(min_row=1, max_row=1) for cell in row):
                    # Print all values under the 'code' column
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
                        for cell in row:
                            if cell.value is not None:
                                print(f"Folder: {os.path.basename(os.path.dirname(file_path))}, File: {os.path.basename(file_path)}, Code: {cell.value}")
                            else:
                                print(f"Folder: {os.path.basename(os.path.dirname(file_path))}, File: {os.path.basename(file_path)}, Code: [Empty]")
                    break
        except Exception as e:
            print(f"Error processing {file_path}: {e}")

# Call the function with the base directory
find_code_column(base_dir)
